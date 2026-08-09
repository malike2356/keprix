"""File limits, hashes, formula safety, and CSV export escaping."""

from __future__ import annotations

import csv
import hashlib
import io
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from keprix.sheet_preprocess.models import SheetInspection

# Align with analytics import ceiling unless caller overrides.
DEFAULT_MAX_BYTES = 10 * 1024 * 1024
DEFAULT_MAX_ROWS = 50_000
DEFAULT_MAX_COLUMNS = 256
DEFAULT_MAX_PROCESSING_SECONDS = 120.0
DEFAULT_MAX_DECOMPRESSED_CELLS = 5_000_000

FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
SUPPORTED_READ_SUFFIXES = frozenset({".csv", ".tsv", ".xlsx", ".xls", ".ods"})
MACRO_SUFFIXES = frozenset({".xlsm"})


class SheetSafetyError(ValueError):
    """Raised when a spreadsheet violates safety policy."""


@dataclass(frozen=True)
class SheetLimits:
    max_bytes: int = DEFAULT_MAX_BYTES
    max_rows: int = DEFAULT_MAX_ROWS
    max_columns: int = DEFAULT_MAX_COLUMNS
    max_processing_seconds: float = DEFAULT_MAX_PROCESSING_SECONDS
    max_decompressed_cells: int = DEFAULT_MAX_DECOMPRESSED_CELLS
    allow_macros: bool = False


def content_hash_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def content_hash_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def looks_like_formula(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.lstrip()
    return bool(text) and text[0] == "="


def escape_csv_cell(value: Any) -> Any:
    """Prefix formula-injection characters so CSV opens as text, not formula."""
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    if value.startswith(FORMULA_INJECTION_PREFIXES):
        return "'" + value
    return value


def write_csv_safe(frame, path: str | Path) -> Path:
    """Write CSV with formula-injection escaping on every cell."""
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    columns = [str(col) for col in frame.columns]
    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        for _, row in frame.iterrows():
            writer.writerow([escape_csv_cell(row[col]) for col in frame.columns])
    return destination


def detect_csv_delimiter(sample: str) -> str:
    """Return an explicit delimiter for reviewable CSV decoding."""
    try:
        dialect = csv.Sniffer().sniff(sample[:8192], delimiters=",\t|;")
        return dialect.delimiter
    except csv.Error:
        if sample.count("\t") > sample.count(","):
            return "\t"
        if sample.count(";") > sample.count(","):
            return ";"
        if sample.count("|") > sample.count(","):
            return "|"
        return ","


def decode_csv_bytes(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = payload.decode(encoding)
            return text, encoding
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace"), "utf-8-replace"


def _count_formula_cells(frame) -> int:
    count = 0
    for column in frame.columns:
        series = frame[column]
        for value in series.tolist():
            if looks_like_formula(value):
                count += 1
    return count


def inspect_workbook_sheets(path: str | Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SheetSafetyError(
            "Excel inspection requires openpyxl; install keprix[analytics]"
        ) from exc
    workbook = load_workbook(filename=str(path), read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def enforce_file_limits(path: str | Path, limits: SheetLimits | None = None) -> SheetInspection:
    """Inspect a file and reject oversized, macro, or unsupported payloads."""
    limits = limits or SheetLimits()
    source = Path(path)
    if not source.is_file():
        raise SheetSafetyError(f"Spreadsheet not found: {source}")
    size = source.stat().st_size
    if size > limits.max_bytes:
        raise SheetSafetyError(
            f"File exceeds size limit ({size} bytes > {limits.max_bytes} bytes)"
        )
    suffix = source.suffix.lower()
    if suffix in MACRO_SUFFIXES and not limits.allow_macros:
        raise SheetSafetyError(
            "Macro-enabled workbooks (.xlsm) are rejected unless allow_macros=True"
        )
    if suffix not in SUPPORTED_READ_SUFFIXES | MACRO_SUFFIXES:
        raise SheetSafetyError(f"Unsupported spreadsheet format: {suffix or '(none)'}")

    warnings: list[str] = []
    worksheets: list[str] = []
    delimiter: str | None = None
    encoding: str | None = None
    flattened = False
    selected: str | int | None = None

    if suffix in {".xlsx", ".xlsm"}:
        try:
            worksheets = inspect_workbook_sheets(source)
        except Exception as exc:
            message = str(exc).lower()
            if "password" in message or "encrypted" in message or "workbook is encrypted" in message:
                raise SheetSafetyError("Encrypted workbooks are not supported") from exc
            raise SheetSafetyError(f"Failed to inspect workbook: {exc}") from exc
        flattened = True
        warnings.append(
            "Output will be a flattened data export; formulas, charts, hidden rows, "
            "and formatting from the original workbook are not preserved"
        )
        if len(worksheets) > 1:
            warnings.append(
                "Multi-sheet workbook requires explicit worksheet selection; "
                f"available sheets: {', '.join(worksheets)}"
            )
        selected = worksheets[0] if len(worksheets) == 1 else None
    elif suffix in {".xls", ".ods"}:
        worksheets = ["sheet1"]
        selected = "sheet1"
        flattened = True
        warnings.append(
            "Legacy spreadsheet will be flattened; formulas are kept as text and never evaluated"
        )
    elif suffix in {".csv", ".tsv"}:
        payload = source.read_bytes()
        text, encoding = decode_csv_bytes(payload)
        delimiter = "\t" if suffix == ".tsv" else detect_csv_delimiter(text)
        worksheets = ["csv"]
        selected = "csv"

    return SheetInspection(
        path=str(source.resolve()),
        content_hash=content_hash_file(source),
        size_bytes=size,
        format=suffix.lstrip("."),
        worksheets=worksheets,
        selected_worksheet=selected,
        delimiter=delimiter,
        encoding=encoding,
        flattened_export=flattened,
        warnings=warnings,
    )


class ProcessingBudget:
    """Wall-clock and rough token budget for batch enrichment."""

    def __init__(
        self,
        *,
        max_processing_seconds: float = DEFAULT_MAX_PROCESSING_SECONDS,
        max_tokens: int = 200_000,
    ):
        self.max_processing_seconds = max_processing_seconds
        self.max_tokens = max_tokens
        self.started_at = time.monotonic()
        self.tokens_used = 0

    def check(self, *, estimated_tokens: int = 0) -> None:
        elapsed = time.monotonic() - self.started_at
        if elapsed > self.max_processing_seconds:
            raise SheetSafetyError(
                f"Processing time limit exceeded ({elapsed:.1f}s > {self.max_processing_seconds}s)"
            )
        if self.tokens_used + estimated_tokens > self.max_tokens:
            raise SheetSafetyError(
                f"Token budget exceeded ({self.tokens_used + estimated_tokens} > {self.max_tokens})"
            )

    def record_tokens(self, tokens: int) -> None:
        self.tokens_used += max(0, int(tokens))


def estimate_batch_tokens(row_count: int, column_count: int) -> int:
    # Rough heuristic for budget gating before model calls.
    return max(1, row_count * column_count * 8)


_PLACEHOLDER_NONEMPTY = frozenset(
    {
        "n/a",
        "na",
        "none",
        "null",
        "tbd",
        "unknown",
        "-",
        "--",
        ".",
    }
)


def is_blank_cell(value: Any) -> bool:
    """Empty means null or whitespace only. Zero, false, formulas, placeholders stay."""
    try:
        import pandas as pd
    except ImportError:
        pd = None  # type: ignore[assignment]
    if pd is not None and bool(pd.isna(value)):
        return True
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return True
        # Formulas and explicit placeholders are user-entered, not empty.
        if looks_like_formula(stripped):
            return False
        if stripped.lower() in _PLACEHOLDER_NONEMPTY:
            return False
        return False
    if value is False or value == 0:
        return False
    return False


def sanitize_loaded_formulas(frame) -> tuple[Any, int]:
    """Keep formula text as opaque strings; never evaluate. Count formula cells."""
    formula_count = 0
    output = frame.copy(deep=True)
    for column in output.columns:
        for idx in range(len(output)):
            value = output.iloc[idx][column]
            if looks_like_formula(value):
                formula_count += 1
                # Store as text so later stages do not treat it as executable.
                output.at[output.index[idx], column] = str(value)
    return output, formula_count


def validate_frame_shape(frame, limits: SheetLimits) -> None:
    rows = len(frame)
    cols = len(frame.columns)
    if rows > limits.max_rows:
        raise SheetSafetyError(f"Row count {rows} exceeds limit {limits.max_rows}")
    if cols > limits.max_columns:
        raise SheetSafetyError(f"Column count {cols} exceeds limit {limits.max_columns}")
    if rows * cols > limits.max_decompressed_cells:
        raise SheetSafetyError(
            f"Cell count {rows * cols} exceeds decompression limit {limits.max_decompressed_cells}"
        )


def load_table_safe(
    path: str | Path,
    *,
    sheet_name: str | int | None = None,
    header_row: int = 0,
    limits: SheetLimits | None = None,
    delimiter: str | None = None,
):
    """Load CSV/TSV/XLSX without evaluating formulas."""
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError(
            "Spreadsheet preprocessing requires pandas; install keprix[analytics]"
        ) from exc

    limits = limits or SheetLimits()
    inspection = enforce_file_limits(path, limits)
    source = Path(path)
    suffix = source.suffix.lower()
    warnings = list(inspection.warnings)

    if suffix == ".csv":
        payload = source.read_bytes()
        text, encoding = decode_csv_bytes(payload)
        chosen = delimiter or inspection.delimiter or detect_csv_delimiter(text)
        frame = pd.read_csv(io.StringIO(text), sep=chosen, header=header_row)
        inspection.delimiter = chosen
        inspection.encoding = encoding
    elif suffix == ".tsv":
        frame = pd.read_csv(source, sep="\t", header=header_row)
        inspection.delimiter = "\t"
    elif suffix in {".xlsx", ".xlsm"}:
        if sheet_name is None:
            if len(inspection.worksheets) > 1:
                raise SheetSafetyError(
                    "Worksheet selection required for multi-sheet workbook: "
                    + ", ".join(inspection.worksheets)
                )
            sheet_name = inspection.worksheets[0] if inspection.worksheets else 0
        # data_only=False: never evaluate formulas; keep formula strings.
        frame = pd.read_excel(
            source,
            sheet_name=sheet_name,
            header=header_row,
            engine="openpyxl",
            dtype=object,
        )
        inspection.selected_worksheet = sheet_name
        inspection.flattened_export = True
    else:
        raise SheetSafetyError(f"Unsupported spreadsheet format: {suffix or '(none)'}")

    validate_frame_shape(frame, limits)
    frame, formula_count = sanitize_loaded_formulas(frame)
    if formula_count:
        warnings.append(
            f"Detected {formula_count} formula cell(s); values kept as text and never evaluated"
        )
    inspection.formula_cells = formula_count
    inspection.row_count = len(frame)
    inspection.column_count = len(frame.columns)
    inspection.header_row = header_row
    inspection.warnings = warnings
    return frame, inspection
