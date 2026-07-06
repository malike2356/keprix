"""Convert uploaded files into tabular text for the analytics workspace."""

from __future__ import annotations

import csv
import io
import json
import re
import tempfile
from pathlib import Path
from typing import Any

MAX_UPLOAD_BYTES = 10 * 1024 * 1024

SUPPORTED_EXTENSIONS = frozenset(
    {
        ".csv",
        ".tsv",
        ".txt",
        ".md",
        ".markdown",
        ".json",
        ".jsonl",
        ".ndjson",
        ".xlsx",
        ".xlsm",
        ".pdf",
        ".docx",
        ".parquet",
        ".sav",
    }
)


class AnalyticsImportError(ValueError):
    """Raised when a file cannot be imported for analytics."""


def supported_analytics_formats() -> list[str]:
    return sorted(ext.lstrip(".") for ext in SUPPORTED_EXTENSIONS)


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _normalize_delimited_text(text: str, *, delimiter: str) -> str:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if delimiter == ",":
        return "\n".join(line for line in lines if line.strip())
    output = io.StringIO()
    writer = csv.writer(output)
    for line in lines:
        if not line.strip():
            continue
        writer.writerow([cell.strip() for cell in line.split(delimiter)])
    return output.getvalue().strip()


def _json_records_to_csv(payload: Any) -> str:
    rows: list[dict[str, Any]]
    if isinstance(payload, list):
        rows = [row for row in payload if isinstance(row, dict)]
    elif isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
                rows = value  # type: ignore[assignment]
                break
        else:
            rows = [payload]
    else:
        raise AnalyticsImportError("JSON must be an array of objects or an object with a records array.")

    if not rows:
        raise AnalyticsImportError("JSON contains no tabular rows.")

    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            key_str = str(key)
            if key_str not in seen:
                seen.add(key_str)
                fieldnames.append(key_str)

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})
    return buffer.getvalue().strip()


def _extract_tabular_from_text(text: str) -> tuple[str, bool]:
    lines = [line.strip() for line in text.replace("\r\n", "\n").split("\n") if line.strip()]
    if not lines:
        raise AnalyticsImportError("No readable text found in file.")

    for delimiter in (",", "\t", "|", ";"):
        counts = [len(line.split(delimiter)) for line in lines[:30]]
        if not counts:
            continue
        if min(counts) >= 2 and max(counts) - min(counts) <= 1:
            if delimiter == ",":
                return "\n".join(lines), True
            return _normalize_delimited_text("\n".join(lines), delimiter=delimiter), True

    if len(lines) >= 2 and all(re.search(r"\d", line) for line in lines[1:3]):
        header = "value"
        return header + "\n" + "\n".join(lines), True

    return "text\n" + "\n".join(lines), False


def _xlsx_to_csv(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AnalyticsImportError(
            "Excel import requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])
            row_count += 1
            if row_count > 5000:
                break
    finally:
        workbook.close()

    data = buffer.getvalue().strip()
    if not data:
        raise AnalyticsImportError("Spreadsheet is empty.")
    return data


def _sav_to_csv(content: bytes) -> str:
    try:
        import pyreadstat
    except ImportError as exc:
        raise AnalyticsImportError(
            "SPSS (.sav) import requires pyreadstat. Install with: pip install pyreadstat"
        ) from exc

    with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        frame, _meta = pyreadstat.read_sav(tmp_path)
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if frame.empty:
        raise AnalyticsImportError("SPSS file has no rows.")
    return frame.to_csv(index=False).strip()


def _parquet_to_csv(content: bytes) -> str:
    try:
        import duckdb
    except ImportError as exc:
        raise AnalyticsImportError(
            "Parquet import requires duckdb. Install with: pip install duckdb"
        ) from exc

    with tempfile.NamedTemporaryFile(suffix=".parquet", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        conn = duckdb.connect()
        frame = conn.execute("SELECT * FROM read_parquet(?)", [tmp_path]).fetchdf()
        conn.close()
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if frame.empty:
        raise AnalyticsImportError("Parquet file has no rows.")
    return frame.to_csv(index=False).strip()


def _pdf_to_text(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise AnalyticsImportError("PDF import requires pypdf.") from exc

    reader = PdfReader(io.BytesIO(content))
    parts = [page.extract_text() or "" for page in reader.pages[:50]]
    text = "\n".join(part.strip() for part in parts if part and part.strip())
    if not text.strip():
        raise AnalyticsImportError("Could not extract text from PDF. Try exporting a CSV from the source.")
    return text


def _docx_to_text(content: bytes) -> str:
    from keprix.documents.parser import parse_docx

    try:
        return parse_docx(content)
    except Exception as exc:
        raise AnalyticsImportError("Could not read Word document.") from exc


def parse_analytics_file(filename: str, content: bytes) -> dict[str, Any]:
    """Parse an uploaded file into CSV-like text for /analytics."""
    if len(content) > MAX_UPLOAD_BYTES:
        raise AnalyticsImportError(f"File is too large. Maximum size is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.")

    suffix = Path(filename or "upload.csv").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        allowed = ", ".join(sorted(supported_analytics_formats()))
        raise AnalyticsImportError(f"Unsupported file type `{suffix or '(none)'}`. Supported: {allowed}")

    source_type = suffix.lstrip(".")
    tabular = True
    message: str | None = None

    if suffix == ".csv":
        data = _decode_text(content)
    elif suffix == ".tsv":
        data = _normalize_delimited_text(_decode_text(content), delimiter="\t")
    elif suffix in {".json", ".jsonl", ".ndjson"}:
        if suffix == ".json":
            payload = json.loads(_decode_text(content))
            data = _json_records_to_csv(payload)
        else:
            rows = []
            for line in _decode_text(content).splitlines():
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
            data = _json_records_to_csv(rows)
        source_type = "json"
    elif suffix in {".xlsx", ".xlsm"}:
        data = _xlsx_to_csv(content)
        source_type = "excel"
    elif suffix == ".parquet":
        data = _parquet_to_csv(content)
        source_type = "parquet"
    elif suffix == ".sav":
        data = _sav_to_csv(content)
        source_type = "spss"
        message = "SPSS file imported. Variable labels are preserved when you use the Research project workflow."
    elif suffix == ".pdf":
        text = _pdf_to_text(content)
        data, tabular = _extract_tabular_from_text(text)
        if not tabular:
            message = "PDF text was imported as a single column. For charts, upload a spreadsheet or CSV if possible."
        source_type = "pdf"
    elif suffix == ".docx":
        text = _docx_to_text(content)
        data, tabular = _extract_tabular_from_text(text)
        if not tabular:
            message = "Word document text was imported as lines. Tables work best when saved as CSV or Excel."
        source_type = "docx"
    elif suffix in {".md", ".markdown"}:
        text = _decode_text(content)
        text = re.sub(r"^#+\s*", "", text, flags=re.MULTILINE)
        data, tabular = _extract_tabular_from_text(text)
        source_type = "markdown"
    else:
        text = _decode_text(content)
        data, tabular = _extract_tabular_from_text(text)
        if not tabular:
            message = "Imported as plain text. Ensure the first row contains column names separated by commas."

    data = data.strip()
    if not data:
        raise AnalyticsImportError("No data could be extracted from this file.")

    row_count = max(0, len([line for line in data.splitlines() if line.strip()]) - 1)
    return {
        "filename": filename,
        "source_type": source_type,
        "data": data,
        "tabular": tabular,
        "row_count": row_count,
        "message": message,
    }
