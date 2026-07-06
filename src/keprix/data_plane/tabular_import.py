"""Excel and SPSS tabular import helpers."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from keprix.data_plane.duckdb_engine import import_csv_dataset


class TabularImportError(ValueError):
    """Raised when a tabular file cannot be imported."""


def _excel_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def _spss_available() -> bool:
    try:
        import pyreadstat  # noqa: F401

        return True
    except ImportError:
        return False


def _write_meta(path: Path, payload: dict[str, Any]) -> str:
    meta_path = path.with_suffix(".meta.json")
    meta_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return str(meta_path)


def import_excel_dataset(
    path: Path,
    *,
    table_name: str = "dataset",
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Import the first (or named) worksheet from an Excel workbook."""
    if not _excel_available():
        raise ImportError("Excel import requires openpyxl; pip install openpyxl")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        sheet_names = list(workbook.sheetnames)
        csv_path = path.with_suffix(".converted.csv")
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
        active_sheet = worksheet.title
    finally:
        workbook.close()

    result = import_csv_dataset(csv_path, table_name=table_name)
    metadata_path = _write_meta(
        path,
        {
            "format": path.suffix.lstrip(".").lower(),
            "sheet": active_sheet,
            "sheet_names": sheet_names,
        },
    )
    return {
        **result,
        "source_format": "excel",
        "sheet": active_sheet,
        "metadata_path": metadata_path,
    }


def import_spss_dataset(path: Path, *, table_name: str = "dataset") -> dict[str, Any]:
    """Import an SPSS .sav file, preserving variable and value labels in metadata."""
    if not _spss_available():
        raise ImportError("SPSS import requires pyreadstat; pip install pyreadstat")
    import pyreadstat

    frame, meta = pyreadstat.read_sav(str(path))
    csv_path = path.with_suffix(".converted.csv")
    frame.to_csv(csv_path, index=False)
    result = import_csv_dataset(csv_path, table_name=table_name)

    value_labels: dict[str, dict[str, str]] = {}
    for column, mapping in (meta.variable_value_labels or {}).items():
        value_labels[str(column)] = {str(key): str(label) for key, label in mapping.items()}

    variable_labels = {
        str(key): str(value) for key, value in (meta.column_names_to_labels or {}).items()
    }
    metadata_path = _write_meta(
        path,
        {
            "format": "sav",
            "variable_labels": variable_labels,
            "value_labels": value_labels,
            "column_names": list(frame.columns),
        },
    )
    return {
        **result,
        "source_format": "spss",
        "variable_labels": variable_labels,
        "value_labels": value_labels,
        "metadata_path": metadata_path,
    }


def supported_tabular_suffixes() -> set[str]:
    base = {".csv", ".tsv", ".parquet"}
    if _excel_available():
        base.update({".xlsx", ".xlsm"})
    if _spss_available():
        base.add(".sav")
    return base
